import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


def mergeFinal():
    # Rutas de los archivos
    ruta_archivo_csv = r"D:\OneDrive - AGROSAVIA - CORPORACION COLOMBIANA DE INVESTIGACION AGROPECUARIA\SampleManager\Desarrollos\InfGenMolecular\CodeInfGenMolecular\Data\Salida\DfFinal\DataFrameFinal.csv"
    ruta_archivo_excel = r"D:\OneDrive - AGROSAVIA - CORPORACION COLOMBIANA DE INVESTIGACION AGROPECUARIA\SampleManager\Desarrollos\InfGenMolecular\CodeInfGenMolecular\Data\GA-F-97 Reporte de resultados laboratorio de servicios una muestra.xlsx"
    ruta_carpeta_salida = r"D:\OneDrive - AGROSAVIA - CORPORACION COLOMBIANA DE INVESTIGACION AGROPECUARIA\SampleManager\Desarrollos\InfGenMolecular\CodeInfGenMolecular\Data\Salida"
    ruta_carpeta_imagenes = r"D:\OneDrive - AGROSAVIA - CORPORACION COLOMBIANA DE INVESTIGACION AGROPECUARIA\SampleManager\Desarrollos\InfGenMolecular\CodeInfGenMolecular\Data\DataGeneticaGenReport\Origen\GraphReport\PNG_Images"

    # Nombre de la hoja de Excel
    nombre_hoja_excel = "GA-F-97"

    # Carga el DataFrame desde el archivo CSV
    df = pd.read_csv(ruta_archivo_csv)

    # Carga el archivo de Excel y la hoja
    libro_excel = load_workbook(ruta_archivo_excel)
    hoja_excel = libro_excel[nombre_hoja_excel]

    # Construye las etiquetas con los nombres de las columnas del DataFrame
    etiquetas = df.columns

    # Realiza el reemplazo y guarda un archivo de Excel por cada fila
    for indice, fila in df.iterrows():
        # Crea una copia del libro de Excel
        libro_copia = load_workbook(ruta_archivo_excel)
        hoja_copia = libro_copia[nombre_hoja_excel]

        # Realiza el reemplazo en la copia de la hoja
        for etiqueta in etiquetas:
            etiqueta_a_reemplazar = f"<<{etiqueta}>>"
            valor_etiqueta = fila[etiqueta]

            for fila_excel in hoja_copia.iter_rows():
                for celda in fila_excel:
                    if celda.value is not None and re.search(
                        etiqueta_a_reemplazar, str(celda.value)
                    ):
                        celda.value = re.sub(
                            etiqueta_a_reemplazar, str(valor_etiqueta), str(celda.value)
                        )

        # Agregar la imagen correspondiente a la celda C68
        idan = fila["idan"]  # Asegúrate de que el nombre de la columna sea correcto
        ruta_imagen = f"{ruta_carpeta_imagenes}\\{idan}.png"

        if os.path.exists(ruta_imagen):
            img = Image(ruta_imagen)
            hoja_copia.add_image(img, "D80")

        # Generar el nombre del archivo de salida
        nombre_archivo_salida = f"INFORME {fila['No de solicitud']} {fila['Código de muestra']} {fila['Nombre de Usuario']} {fila['Now']}.xlsx"
        ruta_archivo_salida = f"{ruta_carpeta_salida}\\{nombre_archivo_salida}"
        libro_copia.save(ruta_archivo_salida)

    print("\n" + "=" * 100)
    print("Se han generado los archivos de Excel.")
    print("=" * 100 + "\n")


# mergeFinal()
